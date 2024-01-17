import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import load_workbook
from lxml import etree


def scrape_ups_and_write_to_excel():
    # Code to go to outlet website and scrape UPS
    rei_url = 'https://www.rei.com/search.html?sort=percentageoff&ir=deals%3AOutlet+Products&r=deals%3AOutlet+Products%3Bbrand%3Apatagonia%7Cmarmot%7Carcteryx%7Ccamelbak%7Ccolumbia%7Cgerber%7Chydro-flask%7Choka%7Cmammut%7Cmerrell%7Cnew-balance%7Cprana%7Csalomon%7Cthe-north-face%7Cunder-armour'
    HEADERS = {
        "User-Agent": "Your-User-Agent-String",
        "Referer": "https://example.com"
    }
    webpage = requests.get(rei_url, headers=HEADERS)
    soup = BeautifulSoup(webpage.content, "html.parser")
    dom = etree.HTML(str(soup))
    ups_data_list = dom.xpath('//*[contains(@href,"product")]')
    if ups_data_list:
        write_to_excel(ups_data_list, "Links")
    else:
        print("Failed to fetch data from outlet website.")


def iterate_excel_and_mark_items():
    # Load Excel workbook
    excel_file_path = 'path_to_desktop/excel_table.xlsx'  # Update with the actual path
    try:
        workbook = load_workbook(excel_file_path)
    except FileNotFoundError:
        print("Excel file not found. Exiting.")
        return

    # Iterate through the sheet
    sheet_name = 'Sheet1'
    if sheet_name not in workbook.sheetnames:
        print(f"Sheet '{sheet_name}' not found in the Excel file. Exiting.")
        return

    sheet = workbook[sheet_name]

    # Check if the 'Marked' column exists, otherwise create it
    marked_column = 'B'
    if marked_column not in sheet[1]:
        sheet[f'{marked_column}1'] = 'Marked'

    # Iterate through each row and get UPC
    # Assuming UPCs are in column A
    for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
        upc = row[0]

        # Code to go to Amazon for each UPC and find rank
        amazon_url = f'https://www.amazon.com/s?k={upc}'
        response = requests.get(amazon_url)

        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')

            # Finding rank logic here (using XPath)
            # Adjust based on actual XPath
            rank_text = soup.find(
                'span', string=lambda text: 'Best Sellers Rank' in text).text
            rank = int(''.join(filter(str.isdigit, rank_text)))

            # Check if rank is less than 100000
            if rank < 100000:
                # Find Amazon price and other necessary details
                amazon_price = float(soup.find('span', class_='a-offscreen').text.replace(
                    '$', '').replace(',', ''))  # Adjust based on actual XPath
                # You may need to find the price on the outlet website
                price = float(input("Enter the price: "))

                # Check if the item is cheap enough
                if amazon_price * 3 <= price:
                    # Mark the item on the Excel table
                    mark_item_on_excel(sheet, upc, marked_column)
        else:
            print(f"Failed to fetch data from Amazon for UPC: {upc}")

    # Save the workbook after processing all UPCs
    workbook.save(excel_file_path)


def clear_excel_sheet(excel_file_path, sheet_name):
    try:
        workbook = load_workbook(excel_file_path)
        sheet = workbook[sheet_name]

        # Delete all rows except the first one (header row)
        sheet.delete_rows(2, sheet.max_row)

        # Save the changes
        workbook.save(excel_file_path)
        print(f"Sheet '{sheet_name}' cleared successfully.")
    except FileNotFoundError:
        print(f"Excel file '{excel_file_path}' not found.")


def write_to_excel(upcList, column_name):
    clear_excel_sheet('/Users/alikozan/Desktop/excel_table.xlsx', 'Sheet')

    # Load or create Excel workbook
    # Update with the actual path

    excel_file_path = '/Users/alikozan/Desktop/excel_table.xlsx'
    try:
        workbook = load_workbook(excel_file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Select or create a sheet
    sheet_name = 'Sheet'
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)

    sheet = workbook[sheet_name]

   # Add column name as the first row
    sheet.append([column_name])
    for upc in upcList:
        sheet.append(["https://www.rei.com" + upc.get('href')])

    # Save the workbook
    workbook.save(excel_file_path)


def mark_item_on_excel(sheet, upc, marked_column):
    # Your logic to mark the item on the Excel table
    print(f"Marking the item with UPC {upc} on the Excel table.")

    # Find the row number based on UPC
    row_number = None
    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] == upc:
            row_number = row[0].row
            break

    if row_number:
        # Mark the 'Marked' column with 'X'
        sheet[f'{marked_column}{row_number}'] = 'X'


# Execute the functions

scrape_ups_and_write_to_excel()
# iterate_excel_and_mark_items()
